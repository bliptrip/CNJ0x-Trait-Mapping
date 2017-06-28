#!/opt/local/bin/python

import sys
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,fills,colors

def parse_args():
    parser = argparse.ArgumentParser(prog=sys.argv[0], description="Compares the dataset in the raw collated file to the Jenny Bolivar curated collated set, returning entries that differ.")
    parser.add_argument('-r', '--raw', action='store', required=True, help="Raw collated spreadsheet file.")
    parser.add_argument('--raw_ws', action='store', required=True, help="Name of raw collated file's worksheet.")
    parser.add_argument('-c', '--curated', action='store', required=True, help="Pre-curated collated spreadsheet file.")
    parser.add_argument('--curated_ws', action='store', required=True, help="")
    parser.add_argument('-o', '--output', action='store', required=True, help="Output excel spreadsheet that will contain the highlighted differences.")
    parser.add_argument('--output_ws', action='store', required=True, help="Output excel workbook's worksheet name.")
    parsed = parser.parse_args(sys.argv[1:])
    return(parsed)

#ACCESSION_TO_INDEX_YEAR={}
ABSOLUTE_YEAR_TO_INDEX_YEAR={2011:1, 2012:2, 2013:3, 2014:4}
RAW_ACCESSION_YEAR_UPRIGHT_TO_ROW={}

def compare_entries(raw_cell, curated_cell):
    string = None
    if(raw_cell.value != curated_cell.value):
        string = "RAW: \"%s\", CURATED: \"%s\""%(str(raw_cell.value), str(curated_cell.value))
    return string

def copy_raw_missing(ws_curated, curated_row, ws_output, output_row, fill):
    for i,cell_string in enumerate(ws_curated['A'+str(curated_row)+':Z'+str(curated_row)][0]):
        ws_output.cell(column=i+1, row=output_row, value=cell_string.value).fill = fill

def copy_prefix(ws_raw, raw_row, ws_output, output_row):
    for i,cell_string in enumerate(ws_raw['A'+str(raw_row)+':D'+str(raw_row)][0]):
        ws_output.cell(column=i+1, row=output_row, value=cell_string.value)

def copy_differences(ws_raw, raw_row, ws_output, output_row, cell_strings, fill):
    for i,cell_string in enumerate(cell_strings):
        if cell_string:
            ws_output.cell(column=i+5, row=output_row, value=cell_string)
            ws_output.cell(column=i+5, row=output_row).fill = fill
        else:
            ws_output.cell(column=i+5, row=output_row, value=ws_raw.cell(column=i+5, row=raw_row).value)


if __name__ == '__main__':
    parsed           = parse_args()
    empty_string     = [None]*22
    wb_raw           = load_workbook(parsed.raw)
    ws_raw           = wb_raw.get_sheet_by_name(parsed.raw_ws)

    #Setup the map of the row/col accession and defined year to an index year
    #initialize the RC_ACCESSION_TO_YEAR dictionary to []
    for row in range(3,ws_raw.max_row+1):
        year            = ws_raw['B'+str(row)].value
        accession       = ws_raw['C'+str(row)].value
        upright         = ws_raw['E'+str(row)].value
        #Commenting the following out as Jenny treated the year index as absolute, not relative to the number of years a given accession was sampled
        #if( accession not in ACCESSION_TO_INDEX_YEAR ):
        #    ACCESSION_TO_INDEX_YEAR[accession] = {}
        #if( year not in ACCESSION_TO_INDEX_YEAR[accession] ):
        #    ACCESSION_TO_INDEX_YEAR[accession][year] = len(ACCESSION_TO_INDEX_YEAR[accession].keys())+1
        #RAW_ACCESSION_YEAR_UPRIGHT_TO_ROW[accession+','+str(ACCESSION_TO_INDEX_YEAR[accession][year])+','+str(upright)] = row
        RAW_ACCESSION_YEAR_UPRIGHT_TO_ROW[accession+','+str(ABSOLUTE_YEAR_TO_INDEX_YEAR[year])+','+str(upright)] = row

    wb_curated       = load_workbook(parsed.curated)
    ws_curated       = wb_curated.get_sheet_by_name(parsed.curated_ws)
        
    wb_output        = load_workbook(parsed.output)
    ws_output        = wb_output.get_sheet_by_name(parsed.output_ws)

    difference_fill  = PatternFill(fgColor=colors.RED,bgColor=colors.RED,fill_type=fills.FILL_SOLID)
    missing_raw_fill = PatternFill(fgColor=colors.YELLOW,bgColor=colors.YELLOW,fill_type=fills.FILL_SOLID)
    ws_output_current_row      = 3

    for row in range(5,ws_curated.max_row+1):
        year            = ws_curated['B'+str(row)].value
        accession       = ws_curated['C'+str(row)].value
        upright         = ws_curated['E'+str(row)].value
        lookup_key      = accession+','+str(year)+','+str(upright)
        if lookup_key in RAW_ACCESSION_YEAR_UPRIGHT_TO_ROW:
            #Compare the values one by one.
            raw_row = RAW_ACCESSION_YEAR_UPRIGHT_TO_ROW[lookup_key]
            raw_cells      = ws_raw['E'+str(raw_row)+':Z'+str(raw_row)][0]
            curated_cells  = ws_curated['E'+str(row)+':Z'+str(row)][0]
            strings = map(compare_entries, raw_cells, curated_cells)
            if( strings.count(None) < 22 ):
                copy_prefix(ws_raw, raw_row, ws_output, ws_output_current_row)
                copy_differences(ws_raw, raw_row, ws_output, ws_output_current_row, strings, difference_fill)
                ws_output_current_row += 1
        else:
            copy_raw_missing(ws_curated, row, ws_output, ws_output_current_row, missing_raw_fill)
            ws_output_current_row += 1
            sys.stderr.write("WARN: Accesssion, Year, Upright %s in curated dataset %s not found in raw dataset %s.\n"%(lookup_key, parsed.curated, parsed.raw))

    wb_output.save(parsed.output)
