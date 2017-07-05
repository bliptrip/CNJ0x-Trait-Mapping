#!/opt/local/bin/python

import re
import sys
import argparse
from openpyxl import load_workbook
from openpyxl import compat

YEAR_ROW_TYPE_TO_FILE_MAP={
"2011,R3,progeny": "2011 Data/BogLower5-R3.xlsx",
"2011,R4,progeny": "2011 Data/BogLower5-R4.xlsx",
"2011,R7,progeny": "2011 Data/BogUpper5-R7.xlsx",
"2011,R8,progeny": "2011 Data/BogUpper5-R8.xlsx",
"2011,R9,progeny": "2011 Data/BogUpper5-R9.xlsx",
"2011,R10,progeny": "2011 Data/BogUpper5-R10.xlsx",
"2011,R44,progeny": "2011 Data/BogUpper5-R44.xlsx",
"2011,R45,progeny": "2011 Data/BogUpper5-R45.xlsx",
"2012,R3,progeny": "2012 Data/BogLower5-R3-1.xlsx",
"2012,R4,progeny": "2012 Data/BogLower5-R4.xlsx",
"2012,R6,parents": "2012 Data/BogLower5-R6 (Parents).xlsx",
"2012,R7,parents": "2012 Data/BogUpper5-R7 (Parents).xlsx",
"2012,R7,progeny": "2012 Data/BogUpper5-R7.xlsx",
"2012,R8,progeny": "2012 Data/BogUpper5-R8.xlsx",
"2012,R9,progeny": "2012 Data/BogUpper5-R9.xlsx",
"2012,R10,progeny": "2012 Data/BogUpper5-R10.xlsx",
"2013,R7,parents": "2013 Data/BogUpper5-R7 (parents).xlsx",
"2013,R7,progeny": "2013 Data/BogUpper5-R7.xlsx",
"2013,R8,progeny": "2013 Data/BogUpper5-R8.xlsx",
"2013,R9,progeny": "2013 Data/BogUpper5-R9.xlsx",
"2013,R10,progeny": "2013 Data/BogUpper5-R10.xlsx",
"2014,R3,progeny": "2014 Data/BogLower5-R3 and R4.xlsx",
"2014,R4,progeny": "2014 Data/BogLower5-R3 and R4.xlsx",
}
    

YEAR_ROW_PARENTAL_TO_WORKBOOK_WORKSHEET_MAP = {}
YEAR_ACCESSION_UPRIGHT_TO_WORKSHEET_ROW_MAP = {}


RE_ROWCOL_HEADER = r'Row/Col:(.*)' 
RE_ROWCOL_ACCESSION = r'\s*R(\d+)\s*C(\d+)\s*' 
RE_ACCESSION_NAME = r'CNJ0[24]-\d+-\d+'
RE_UPRIGHT = r'\s*(\d+)\s*' 
#The following was introduced to cross-check the validity of the Row designator in the input files, as there were cases of incorrect rows."
RE_OUTPUT_FILE_ROW = r'Bog(Upper)|(Lower)5-R(\d+).xlsx'
STATE_INIT = 0
STATE_COPY = 1
state = STATE_INIT

def parse_args():
    parser = argparse.ArgumentParser(prog=sys.argv[0], description="Copies entries in a collated curated file to the original dataset files (because I failed to have us edit the original files when hand curating datasets).")
    parser.add_argument('-c', '--curate', action='store', required=True, help="Hand curated datafile in collated form.")
    parser.add_argument('--curate_ws', action='store', required=True, help="Name of curated datafile's worksheet.")
    parser.add_argument('--data_prefix', action='store', required=True, help='Data directory prefix to properly access the datafiles.')
    parsed = parser.parse_args(sys.argv[1:])
    return(parsed)

def copy_row(ws_dest, dest_row, ws_source, source_row):
    for i,source_cell in enumerate(ws_source['E'+str(source_row)+':AB'+str(source_row)][0]):
        ws_dest.cell(column=i+1,row=dest_row,value=source_cell.value)

if __name__ == '__main__':
    parsed           = parse_args()

    wb_curate         = load_workbook(parsed.curate)
    ws_curate         = wb_curate.get_sheet_by_name(parsed.curate_ws)
    curate_max_row     = ws_curate.max_row

    re_rowcol_accession = re.compile(RE_ROWCOL_ACCESSION, re.I)
    re_accession_name    = re.compile(RE_ACCESSION_NAME, re.I)
    re_upright            = re.compile(RE_UPRIGHT, re.I)

    #Scan the curate file to determine which original dataset files to open
    year_row_type_list = []    
    for row in range(3, curate_max_row+1):
        year = str(ws_curate['B'+str(row)].value).strip()
        #Determine the row
        accession = ws_curate['C'+str(row)].value.strip()
        re_rowcol_accession_match = re_rowcol_accession.match(accession)
        if( re_rowcol_accession_match ):
                accession_row = 'R'+str(re_rowcol_accession_match.group(1))
        else:
                sys.stderr.write("WARN: Unable to find a row designator in accession %s (spreadsheet row %d)\n" % (accession,row))
                continue
        #Determine the type (progeny or parents)
        accession_name = ws_curate['D'+str(row)].value.strip()
        re_accession_name_match = re_accession_name.match(accession_name)
        if (re_accession_name_match):
                accession_type = "progeny"
        else:
                accession_type = "parents"
        lookup_key = "%s,%s,%s" % (year,accession_row,accession_type)
        if lookup_key not in YEAR_ROW_PARENTAL_TO_WORKBOOK_WORKSHEET_MAP:
                if lookup_key in YEAR_ROW_TYPE_TO_FILE_MAP:
                        wb_file = YEAR_ROW_TYPE_TO_FILE_MAP[lookup_key]
                        wb_out  = load_workbook(parsed.data_prefix+"/"+wb_file)
                        ws_out  = wb_out.get_sheet_by_name("Sheet1")
                        YEAR_ROW_PARENTAL_TO_WORKBOOK_WORKSHEET_MAP[lookup_key] = (wb_out,ws_out)
                else:
                        sys.stderr.write("ERR: Failed to find lookup key \"%s\" in file map (spreadsheet row %d)\n" % (lookup_key,row))

    #Subsequently, open each needed curate file and build a map of year accessions, uprights to the appropriate row number
    re_expected_row = re.compile(RE_OUTPUT_FILE_ROW, re.I)

    for key,item in YEAR_ROW_PARENTAL_TO_WORKBOOK_WORKSHEET_MAP.items():
        year            = key.split(',')[0]
        output_filename = YEAR_ROW_TYPE_TO_FILE_MAP[key]
        (wb_out,ws_out) = YEAR_ROW_PARENTAL_TO_WORKBOOK_WORKSHEET_MAP[key]

        expected_row = 0
        re_expected_row_match  = re_expected_row.match(output_filename)
        if(re_expected_row_match):
                expected_row = re_expected_row_match.group(3)

        re_header_compile = re.compile(RE_ROWCOL_HEADER, re.I)
        re_accession_compile = re.compile(RE_ROWCOL_ACCESSION, re.I)
        state = STATE_INIT
        for row in range(1, ws_out.max_row + 1):
            header = ws_out['A'+str(row)].value
            re_header_match = re_header_compile.match(str(header))
            if( re_header_match ):
                pre_accession = re_header_match.group(1)
                re_accession_match = re_accession_compile.match(str(pre_accession))
                if( not re_accession_match ):
                    pre_accession = ws_out['C'+str(row)].value
                    re_accession_match = re_accession_compile.match(str(pre_accession))
                if( re_accession_match ):
                    if(expected_row and (re_accession_match.group(1) != expected_row)):
                        sys.stderr.write("WARN: Expected row identifier: %s, Actual row identifier: %s\n" % (expected_row, re_accession_match.group(1)))
                    accession = 'R'+str(re_accession_match.group(1))+'C'+str(re_accession_match.group(2))
                    state = STATE_COPY
                    current_upright_id = 1
            elif state == STATE_COPY:
                    re_upright_match = re_upright.match(str(header))
                    if(re_upright_match):
                        upright_id = int(re_upright_match.group(1))
                        if( upright_id == current_upright_id ):
                            lookup_key = "%s,%s,%d" % (year,accession,upright_id)
                            YEAR_ACCESSION_UPRIGHT_TO_WORKSHEET_ROW_MAP[lookup_key] = row
                        else:
                            syst.stderr.write("WARN: Potentially invalid upright id: %d, in file %s, row %d\n" % (upright_id,output_filename,row))
                        current_upright_id = upright_id + 1
                    else:
                        state = STATE_INIT

    #Now rescan curated file, using year, accession, and upright to lookup destination spreadsheet and row to copy to
    for row in range(3, curate_max_row+1):
        #Determine the year
        year = str(ws_curate['B'+str(row)].value).strip()
        #Determine the accession
        accession = ws_curate['C'+str(row)].value.strip()
        re_rowcol_accession_match = re_rowcol_accession.match(accession)
        if( re_rowcol_accession_match ):
                accession_row = 'R'+str(re_rowcol_accession_match.group(1))
        else:
                sys.stderr.write("WARN: Unable to find a row designator in accession %s (spreadsheet row %d)\n" % (accession,row))
                continue
        #Determine the type (progeny or parents)
        accession_name = ws_curate['D'+str(row)].value.strip()
        re_accession_name_match = re_accession_name.match(accession_name)
        if (re_accession_name_match):
                accession_type = "progeny"
        else:
                accession_type = "parents"
        #Determine the upright id
        upright_id = str(ws_curate['E'+str(row)].value).strip()
        #Generate lookup keys
        lookup_key_workbook = "%s,%s,%s" % (year,accession_row,accession_type)
        lookup_key_accession_row = "%s,%s,%s" % (year,accession,upright_id)

        (wb_out,ws_out) = YEAR_ROW_PARENTAL_TO_WORKBOOK_WORKSHEET_MAP[lookup_key_workbook]
        output_row = YEAR_ACCESSION_UPRIGHT_TO_WORKSHEET_ROW_MAP[lookup_key_accession_row]
        copy_row(ws_out, output_row, ws_curate, row)
    
    #Now loop through the open workbooks and save them
    for key,item in YEAR_ROW_PARENTAL_TO_WORKBOOK_WORKSHEET_MAP.items():
        output_filename = parsed.data_prefix+'/'+YEAR_ROW_TYPE_TO_FILE_MAP[key]
        wb_out = item[0]
        wb_out.save(output_filename)

