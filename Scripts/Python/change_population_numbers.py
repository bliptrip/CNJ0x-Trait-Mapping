#!/opt/local/bin/python

import re
import sys
import argparse
from openpyxl import load_workbook
from openpyxl import compat

RE_POP1 = "R[346]\s*C\d+"
RE_POP2 = "R([789]|10|44|45)\s*C\d+"

def parse_args():
    parser = argparse.ArgumentParser(prog=sys.argv[0], description="Fill in the population numbers in the excel spreadsheet.")
    parser.add_argument('-i', '--input', action='store', required=True, help="Input excel file.")
    parsed = parser.parse_args(sys.argv[1:])
    return(parsed)

if __name__ == '__main__':
    parsed           = parse_args()
    wb               = load_workbook(parsed.input)
    #Select the active worksheet
    ws               = wb.active
    ws.title         = "Combined CNJ Upright Data"
    re_p1            = re.compile(RE_POP1)
    re_p2            = re.compile(RE_POP2)
    for row in range(2,ws.max_row+1):
        accession = ws['C'+str(row)].value.strip()
        ws['C'+str(row)].value = accession
        if re_p1.match(accession):
            ws['A'+str(row)].value = 1
        elif re_p2.match(accession):
            ws['A'+str(row)].value = 2
        else:
            sys.stderr.write("WARN: Invalid accession: %s\n" % (accession))
    wb.save(filename = parsed.input + "copy.xls")
