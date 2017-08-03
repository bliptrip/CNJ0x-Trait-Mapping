#!/opt/local/bin/python

import sys
import argparse
import re
from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill,fills,colors

RE_ROWCOL_HEADER = r'Row/Col' 

def parse_args():
    parser = argparse.ArgumentParser(prog=sys.argv[0], description="Does a raw diff of two spreadsheets, outputting the differences to a file with differences highlighted in red.")
    parser.add_argument('-1', '--file_1', action='store', required=True, help="Spreadsheet 1")
    parser.add_argument('--file_1_ws', action='store', required=True, help="")
    parser.add_argument('-2', '--file_2', action='store', required=True, help="Spreadsheet 2")
    parser.add_argument('--file_2_ws', action='store', required=True, help="Spreadsheet 2")
    parser.add_argument('-o', '--output', action='store', required=True, help="Output excel spreadsheet that will contain the highlighted differences.")
    parser.add_argument('--output_ws', action='store', required=True, help="Output excel workbook's worksheet name.")
    parsed = parser.parse_args(sys.argv[1:])
    return(parsed)

def compare_entries(raw_cell, curated_cell):
    string = None
    if(raw_cell.value != curated_cell.value):
        string = "RAW: \"%s\", CURATED: \"%s\""%(str(raw_cell.value), str(curated_cell.value))
    return string

def copy_cells(ws_output, output_row, ws_source_cells):
    for i,cell in enumerate(ws_source_cells):
        ws_output.cell(column=i+1, row=output_row, value=cell.value)

def copy_differences(ws_file, file_row, ws_output, output_row, cell_strings, fill):
    for i,cell_string in enumerate(cell_strings):
        if cell_string:
            ws_output.cell(column=i+1, row=output_row, value=cell_string)
            ws_output.cell(column=i+1, row=output_row).fill = fill
        else:
            ws_output.cell(column=i+1, row=output_row, value=ws_file.cell(column=i+1, row=file_row).value)

if __name__ == '__main__':
    parsed           = parse_args()

    wb_file_1        = load_workbook(parsed.file_1)
    ws_file_1        = wb_file_1.get_sheet_by_name(parsed.file_1_ws)

    wb_file_2        = load_workbook(parsed.file_2)
    ws_file_2        = wb_file_2.get_sheet_by_name(parsed.file_2_ws)

    wb_output        = Workbook()
    ws_output        = wb_output.active

    difference_fill        = PatternFill(fgColor=colors.RED,bgColor=colors.RED,fill_type=fills.FILL_SOLID)
    re_header_compile      = re.compile(RE_ROWCOL_HEADER, re.I)
    ws_output_current_row  = 1
    total_errors_count     = 0
    total_genotypes        = 0

    
    #Setup the map of the row/col accession and defined year to an index year
    #initialize the RC_ACCESSION_TO_YEAR dictionary to []
    max_row = max(ws_file_1.max_row,ws_file_2.max_row)
    for row in range(1,max_row+1):
        file_1_cells   = ws_file_1['A'+str(row)+':V'+str(row)][0]
        file_2_cells   = ws_file_2['A'+str(row)+':V'+str(row)][0]
        if re_header_compile.match(str(file_2_cells[0].value)):
            total_genotypes += 1
        expected_column_count = len(file_1_cells)
        strings = map(compare_entries, file_1_cells, file_2_cells)
        total_errors_count += (len(strings) - strings.count(None))
        if( strings.count(None) < expected_column_count ):
            copy_differences(ws_file_1, row, ws_output, ws_output_current_row, strings, difference_fill)
            ws_output_current_row += 1
        else:
            copy_cells(ws_output, ws_output_current_row, file_2_cells)
            ws_output_current_row += 1

    error_count_row = 2
    ws_output.cell(column=1, row=error_count_row, value="No. Error Phenotype").fill = difference_fill
    ws_output.cell(column=2, row=error_count_row, value="%d"%(total_errors_count)).fill = difference_fill
    approximate_num_phenotypes = total_genotypes * 10 * 21 #approximately 10 uprights per genotype, 21 phenotypic variables measured per upright
    ws_output.cell(column=3, row=error_count_row, value="Approx No. Total Phenotypes").fill = difference_fill
    ws_output.cell(column=4, row=error_count_row, value="%d"%(approximate_num_phenotypes)).fill = difference_fill
    ws_output.cell(column=5, row=error_count_row, value="Percent Error").fill = difference_fill
    ws_output.cell(column=6, row=error_count_row, value="%.2f"%((total_errors_count*100.0)/approximate_num_phenotypes)).fill = difference_fill

    wb_output.save(parsed.output)
