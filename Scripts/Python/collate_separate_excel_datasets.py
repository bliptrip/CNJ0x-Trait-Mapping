#!/opt/local/bin/python

import re
import sys
import argparse
from openpyxl import load_workbook
from openpyxl import compat

RE_ROWCOL_HEADER = r'Row/Col:(.*)' 
RE_ROWCOL_ACCESSION = r'\s*R(\d+)\s*C(\d+)\s*' 
#The following was introduced to cross-check the validity of the Row designator in the input files, as there were cases of incorrect rows."
RE_INPUT_FILE_ROW = r'Bog(Upper)|(Lower)5-R(\d+).xlsx'
STATE_INIT = 0
STATE_COPY = 1
state_index = 1
state = STATE_INIT

def parse_args():
    parser = argparse.ArgumentParser(prog=sys.argv[0], description="Collate a specified input NJ cranberry dataset worksheet an existing combined workbook.")
    parser.add_argument('-i', '--input', action='store', required=True, help="Input excel workbook to pull from.")
    parser.add_argument('--input_ws', action='store', required=True, help="Input excel workbook's worksheet name.")
    parser.add_argument('-o', '--output', action='store', required=True, help="Output excel workbook concatenate collated data to.")
    parser.add_argument('--output_ws', action='store', required=True, help="Output excel workbook's worksheet name.")
    parser.add_argument('-y', '--year', action='store', type=int, required=True, help="Output excel workbook year.")
    parser.add_argument('-p', '--population', action='store', required=True, help="Population identifier")
    parsed = parser.parse_args(sys.argv[1:])
    return(parsed)

def collate_copy_row(ws_dest, dest_row, dest_column, source_cells):
    for source_cell in source_cells:
        ws_dest.cell(column=dest_column,row=dest_row, value=source_cell.value)
        dest_column += 1

if __name__ == '__main__':
    parsed           = parse_args()

    wb_input         = load_workbook(parsed.input)
    ws_input         = wb_input.get_sheet_by_name(parsed.input_ws)
    wb_output        = load_workbook(parsed.output)
    ws_output        = wb_output.get_sheet_by_name(parsed.output_ws)
    ws_output_curr_row = ws_output.max_row+1
    accession        = "None"
    state            = STATE_INIT
    expected_row     = 0

    #Extract the expected row designator from the file."
    re_expected_row_compile = re.compile(RE_INPUT_FILE_ROW, re.I)
    re_expected_row_match  = re_expected_row_compile.match(parsed.input)
    if(re_expected_row_match):
        expected_row = re_expected_row_match.group(3)

    re_header_compile = re.compile(RE_ROWCOL_HEADER, re.I)
    re_accession_compile = re.compile(RE_ROWCOL_ACCESSION, re.I)
    for row in range(1, ws_input.max_row + 1):
        header = ws_input['A'+str(row)].value
        re_header_match = re_header_compile.match(str(header))
        if( re_header_match ):
            pre_accession = re_header_match.group(1)
            re_accession_match = re_accession_compile.match(str(pre_accession))
            if( not re_accession_match ):
                pre_accession = ws_input['C'+str(row)].value
                re_accession_match = re_accession_compile.match(str(pre_accession))
            if( re_accession_match ):
                if(expected_row and (re_accession_match.group(1) != expected_row)):
                    sys.stderr.write("WARN: Expected row identifier: %s, Actual row identifier: %s\n" % (expected_row, re_accession_match.group(1)))
                accession = 'R'+str(re_accession_match.group(1))+'C'+str(re_accession_match.group(2))
                state_index = 1
                state = STATE_COPY
        elif state == STATE_COPY:
            if( state_index <= 10 ):
                ws_output.cell(row=ws_output_curr_row,column=1, value=parsed.population)
                ws_output.cell(row=ws_output_curr_row,column=2, value=parsed.year)
                ws_output.cell(row=ws_output_curr_row,column=3, value=accession)
                ws_output.cell(row=ws_output_curr_row,column=5, value=state_index)
                upright_entries = ws_input['B'+str(row)+":"+'Z'+str(row)][0]
                collate_copy_row(ws_dest=ws_output, dest_row=ws_output_curr_row, dest_column=6, source_cells=upright_entries)
                ws_output_curr_row += 1
                state_index += 1
            else:
                state = STATE_INIT
    wb_output.save(filename=parsed.output)
