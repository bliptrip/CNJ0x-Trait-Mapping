#!/opt/local/bin/python

import sys
import argparse
from openpyxl import Workbook,load_workbook
from numpy.random import randint


def parse_args():
    parser = argparse.ArgumentParser(prog=sys.argv[0], description="Randomly select a subset of entries from the cranberry phenotyping data.")
    parser.add_argument('-i', '--input', action='store', required=True, help="Input excel workbook to select subset from.")
    parser.add_argument('--input_ws', action='store', required=True, help="Input excel workbook's worksheet name.")
    parser.add_argument('-o', '--output', action='store', required=True, help="Output excel workbook that contained subset.")
    parser.add_argument('--output_ws', action='store', required=True, help="Output excel workbook's worksheet name.")
    parser.add_argument('-p', '--percentage', action='store', type=int, required=True, help="The percentage of input entries to select a random sample from.")
    parsed = parser.parse_args(sys.argv[1:])
    return(parsed)


if __name__ == '__main__':
    parsed           = parse_args()

    wb_input         = load_workbook(parsed.input)
    ws_input         = wb_input.get_sheet_by_name(parsed.input_ws)
    wb_output        = Workbook()
    ws_output        = wb_output.create_sheet(title=parsed.output_ws)
    ws_input_num_entries = ws_input.max_row-2
    ws_output_sample_num_entries = int((float(parsed.percentage)/100.00)*ws_input_num_entries)
    subsample_indices = randint(low=1, high=ws_input_num_entries+1, size=ws_output_sample_num_entries)

    def copy_to_new_workbook(input_row):
        upright_entries = ws_input['A'+str(input_row)+":"+'AB'+str(input_row)][0]
        output_row=ws_output.max_row+1
        for (i,upright_entry) in enumerate(upright_entries):
            ws_output.cell(column=i+1,row=output_row, value=upright_entry.value)

    copy_to_new_workbook(1)
    copy_to_new_workbook(2)
    map(copy_to_new_workbook, subsample_indices+2)

    wb_output.save(filename=parsed.output)
