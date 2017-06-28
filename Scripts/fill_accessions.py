#!/opt/local/bin/python

import sys
import argparse
from openpyxl import load_workbook

def parse_args():
    parser = argparse.ArgumentParser(prog=sys.argv[0], description="Fill in the NJ accession id's as derived from a row/column designator.")
    parser.add_argument('-i', '--input', action='store', required=True, help="Input excel workbook to pull NJ accession ID's from.")
    parser.add_argument('--input_ws', action='store', required=True, help="Input excel workbook's worksheet name with the map.")
    parser.add_argument('-o', '--output', action='store', required=True, help="Output excel workbook with collated data to add NJ accessions to.")
    parser.add_argument('--output_ws', action='store', required=True, help="Output excel workbook's worksheet name.")
    parsed = parser.parse_args(sys.argv[1:])
    return(parsed)

if __name__ == '__main__':
    parsed           = parse_args()

    wb_input         = load_workbook(parsed.input)
    ws_input         = wb_input.get_sheet_by_name(parsed.input_ws)
    wb_output        = load_workbook(parsed.output)
    ws_output        = wb_output.get_sheet_by_name(parsed.output_ws)
    #initialize the map dictionary to empty
    cranmap          = {}

    #First loop through the map file and generate a dictionary that maps the Row-Column
    for row in range(2,ws_input.max_row+1):
        accession       = 'R'+str(ws_input['A'+str(row)].value)+'C'+str(ws_input['B'+str(row)].value)
        accession_name  = ws_input['C'+str(row)].value
        cranmap[accession] = accession_name

    #Now go through the destination file and insert the correct accession name for each accession.
    for row in range(3,ws_output.max_row+1):
        accession = ws_output['C'+str(row)].value.strip()
        accession_name = cranmap[accession]
        ws_output['D'+str(row)].value = accession_name
    wb_output.save(filename=parsed.output)
